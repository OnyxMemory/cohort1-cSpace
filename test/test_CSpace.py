import unittest
import openpyxl
from CSpace import CSpace
from Client import Client


class TestOopMajorCSpace(unittest.TestCase):
    def setUp(self):
        workbook = openpyxl.Workbook()

        self.month_sheet = workbook.active
        self.month_sheet.title = '2018-08'

        month = (
            (None, None, 'First Floor Hall', 'Second Floor Hall', 'Desk 3rd floor 1', 'Desk 3rd floor 7', 'RGO Treehouse (South)', 'RGO Treehouse (North)'),
            ('2018-08-01', 'Wed', '', '', '', 'James', 'Don', ''),
            ('2018-08-02', 'Thu', 'Carrie', 'James', '', 'Don', '', ''),
            ('2018-08-03', 'Fri', 'Carrie', 'Don', 'Don', '', '', ''),
            ('2018-08-04', 'Sat', 'James', '', '', 'James', 'Don', 'Carrie'),
            ('2018-08-05', 'Wed', '', '', '', '', '', ''),
            ('2018-08-06', 'Wed', '', '', '', '', '', ''),
            ('2018-08-07', 'Wed', '', '', '', '', '', ''),
            ('2018-08-08', 'Wed', '', '', '', '', '', ''),
            ('2018-08-09', 'Wed', '', '', '', '', '', ''),
            ('2018-08-10', 'Wed', '', '', '', '', '', ''),
            ('2018-08-11', 'Wed', '', '', '', '', '', ''),
            ('2018-08-12', 'Wed', '', '', '', '', '', ''),
            ('2018-08-13', 'Wed', '', '', '', '', '', ''),
            ('2018-08-14', 'Wed', '', '', '', '', '', ''),
            ('2018-08-15', 'Wed', '', '', '', '', '', ''),
            ('2018-08-16', 'Wed', '', '', '', '', '', ''),
            ('2018-08-17', 'Wed', '', '', '', '', '', ''),
            ('2018-08-18', 'Wed', '', '', '', '', '', ''),
            ('2018-08-19', 'Wed', '', '', '', '', '', ''),
            ('2018-08-20', 'Wed', '', '', '', '', '', ''),
            ('2018-08-21', 'Wed', '', '', '', '', '', ''),
            ('2018-08-22', 'Wed', '', '', '', '', '', ''),
            ('2018-08-23', 'Wed', '', '', '', '', '', ''),
            ('2018-08-24', 'Wed', '', '', '', '', '', ''),
            ('2018-08-25', 'Wed', '', '', '', '', '', ''),
            ('2018-08-26', 'Wed', '', '', '', '', '', ''),
            ('2018-08-27', 'Wed', '', '', '', '', '', ''),
            ('2018-08-28', 'Wed', '', '', '', '', '', ''),
            ('2018-08-29', 'Wed', '', '', '', '', '', ''),
            ('2018-08-30', 'Wed', '', '', '', '', '', ''),
            ('2018-08-31', 'Wed', '', '', '', '', '', ''),
        )

        for row in month:
            self.month_sheet.append(row)

        self.client_sheet = workbook.create_sheet('Clients')
        self.rate_sheet = workbook.create_sheet('Rates')
        self.room_sheet = workbook.create_sheet('Facilities')

        clients = (
            ('Name', 'Email', 'Phone', 'Payment Method', 'Payment Status', 'Issues', 'Notes'),
            ('Carrie Cordon', 'carrie.cordon@gmail.com', '403-555-6666', 'VISA', 'PAID', 'Cranky when mad',
             'Pays on time'),
            ('James Jordan', 'james.jordan@gmail.com', '403-444-6666', 'VISA', 'LATE', '', 'Has not paid on time once'),
            ('Don Dunkirk', 'don.dunkirk@gmail.com', '403-555-7777', 'MASTERCARD', 'PAID', 'Never is happy',
             'Complains too much')
        )

        for row in clients:
            self.client_sheet.append(row)

        rates = (
            ('Location', 'Type', 'Credits'),
            ('King Edward', 'Desk', 1),
            ('King Edward', 'Gallery', 5),
            ('King Edward', 'Meeting', 3)
        )

        for row in rates:
            self.rate_sheet.append(row)

        rooms = (
            ('Room Name', 'Type', 'Location'),
            ('First Floor Hall', 'Gallery', 'King Edward', '', ''),
            ('Second Floor Hall', 'Gallery', 'King Edward', '', ''),
            ('Desk 3rd floor 1', 'Desk', 'King Edward', '', ''),
            ('Desk 3rd floor 7', 'Desk', 'King Edward', '', ''),
            ('RGO Treehouse (South)', 'Meeting', 'King Edward', '', ''),
            ('RGO Treehouse (North)', 'Meeting', 'King Edward', '', '')
        )

        for row in rooms:
            self.room_sheet.append(row)

        self.client_data = \
                    ('Carrie Cordon',
                     'carrie.cordon@gmail.com',
                     '404-444-1111',
                     'VISA',
                     'PAID',
                     'Doesn\'t pay on time',
                     'Give notice to pay.')

        self.cspace = CSpace(workbook)

        self.booking_dict = self.cspace.extract_bookings(self.month_sheet)
        self.clients_array = self.cspace.extract_data(self.client_sheet)
        self.room_array = self.cspace.extract_data(self.room_sheet)

    def test_add_client(self):
        test_client = Client(self.client_data)

        self.assertEqual(0, len(self.cspace.clients))
        self.cspace.add_client(test_client)
        self.assertEqual(1, len(self.cspace.clients))

    def test_extract_rates(self):
        rates_dict = {
            'Desk': 1,
            'Gallery': 5,
            'Meeting': 3
        }
        self.assertEqual(rates_dict, CSpace.extract_rates(self.rate_sheet))

    def test_extract_bookings(self):
        booking_dict = {
            'First Floor Hall': ['', 'Carrie', 'Carrie', 'James'],
            'Second Floor Hall': ['', 'James', 'Don', ''],
            'Desk 3rd floor 1': ['', '', 'Don', ''],
            'Desk 3rd floor 7': ['James', 'Don', '', 'James'],
            'RGO Treehouse (South)': ['Don', '', '', 'Don'],
            'RGO Treehouse (North)': ['', '', '', 'Carrie']
        }
        for k,v in booking_dict.items():
            for i in range(27):
                v.append('')

        self.assertEqual(booking_dict, CSpace.extract_bookings(self.month_sheet))

    def test_extract_clients(self):
        clients_array = [
            ['Carrie Cordon', 'carrie.cordon@gmail.com', '403-555-6666', 'VISA', 'PAID', 'Cranky when mad',
             'Pays on time'],
            ['James Jordan', 'james.jordan@gmail.com', '403-444-6666', 'VISA', 'LATE', '', 'Has not paid on time once'],
            ['Don Dunkirk', 'don.dunkirk@gmail.com', '403-555-7777', 'MASTERCARD', 'PAID', 'Never is happy',
             'Complains too much']
        ]

        self.assertEqual(clients_array, CSpace.extract_data(self.client_sheet))

    def test_add_rooms_from_array(self):
        self.assertEqual(0, len(self.cspace.rooms))
        self.cspace.add_rooms_from_array(self.room_array, self.rate_sheet)
        self.assertEqual(6, len(self.cspace.rooms))

    def test_add_clients_from_array(self):
        self.assertEqual(0, len(self.cspace.clients))
        self.cspace.add_clients_from_array(self.clients_array)
        self.assertEqual(3, len(self.cspace.clients))

    def test_update_credits(self):
        self.cspace.add_clients_from_array(self.clients_array)
        self.cspace.add_rooms_from_array(self.room_array, self.rate_sheet)

        self.assertEqual(0, self.cspace.clients['Carrie'].credits)
        self.assertEqual(0, self.cspace.clients['James'].credits)
        self.cspace.update_credits(self.booking_dict)
        self.assertEqual(13, self.cspace.clients['Carrie'].credits)
        self.assertEqual(12, self.cspace.clients['James'].credits)

    def test_run(self):
        output = [
            ('Carrie', 'Cordon', 13),
            ('James', 'Jordan', 12),
            ('Don', 'Dunkirk', 13)
        ]

        self.assertEqual(output, self.cspace.run('2018-08'))

    def test_populate_client_bookings(self):
        self.cspace.add_clients_from_array(self.clients_array)
        test_bookings = {'2018-08':[(2,'First Floor Hall'), (3,'First Floor Hall'), (4,'RGO Treehouse (North)')]}

        self.assertEqual({},self.cspace.clients['Carrie'].bookings)
        self.cspace.populate_client_bookings()
        self.assertEqual(test_bookings,self.cspace.clients['Carrie'].bookings)
